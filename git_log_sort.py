import datetime
import whichTeam

import openpyxl
workbook          = openpyxl.load_workbook('gitLog.xlsx')
history_workbook  = openpyxl.load_workbook('gitLogHistory.xlsx')


def wt_xlsx(author_name, git_log_split, date_str, team):
 
    global workbook, history_workbook
    print(team + "增加一個 commit log \n")
    
    git_log_merge=''
    for text in git_log_split:
        git_log_merge+=text

    commit_date = datetime.datetime.strptime(date_str, "%b %d %H:%M:%S %Y %z")

    sheet  = workbook.get_sheet_by_name(team)
    
    startRow=sheet.max_row+1
    sheet['A' + str(startRow)] = author_name
    sheet['B' + str(startRow)] = git_log_merge
    sheet['C' + str(startRow)] = commit_date

    history_sheet = history_workbook.get_sheet_by_name(team)
    
    historyStartRow=history_sheet.max_row+1
    history_sheet['A' + str(historyStartRow)] = author_name
    history_sheet['B' + str(historyStartRow)] = git_log_merge
    history_sheet['C' + str(historyStartRow)] = commit_date

#note1 是 "非"merge 的 git log
def is_note_head1(lt, i):
    if 'commit' == lt[i][0:6] and 'Author:' == lt[i + 1][0:7] and 'Date:' == lt[i + 2][0:5]:
        return True
    return False

#將commit, date, author, release note 整合成一筆資料, release note要另外處理
def add_note(git_log_list, note_head_flag):
    for j in range(len(note_head_flag)):
        print("第" + str(j) + "筆 commit")
        
        commit_flag = note_head_flag[j]
        author_flag = note_head_flag[j] + 1
        date_flag   = note_head_flag[j] + 2
        
        note_text = []
        
        note_text.append(git_log_list[commit_flag])

        note_text.append(git_log_list[author_flag])
        memberName=git_log_list[author_flag][8:]
        print("作者是 " + memberName.split("<",1)[0])        
        team=whichTeam.whichTeam(memberName)
        print("屬於 " + team + " team")
        
        note_text.append(git_log_list[date_flag])
        #release note 使用別的function來整合
        release_note_text = make_release_note(j, note_head_flag, is_note_head2, git_log_list)        
        note_text.append(release_note_text)#加入 release note

        wt_xlsx(memberName, note_text, git_log_list[date_flag][12:-1],team)


#note2 是 merge 的 git log
def is_note_head2(lt, i):
    if 'commit' == lt[i][0:6] and 'Merge:' == lt[i+1][0:6] and  'Author:' == lt[i+2][0:7] and 'Date:' in lt[i+3][0:5]:
        return True
    return False

#將commit, merge, date, author, release note 整合成一筆資料, release note要另外處理
def add_note2(git_log_list, note_head_flag):
    
    for j in range(len(note_head_flag)):
        print("第" + str(j) + "筆 merge commit")
        
        commit_flag = note_head_flag[j]
        merge_flag  = note_head_flag[j] + 1
        author_flag = note_head_flag[j] + 2
        date_flag   = note_head_flag[j] + 3
        
        note_text = []
        
        note_text.append(git_log_list[commit_flag])
        
        note_text.append(git_log_list[merge_flag])

        note_text.append(git_log_list[author_flag])
        memberName=git_log_list[author_flag][8:]
        print("作者是 " + memberName.split("<",1)[0])        
        team=whichTeam.whichTeam(memberName)
        print("屬於 " + team + " team")
        
        note_text.append(git_log_list[date_flag])
        #release note 使用別的function來整合
        release_note_text = make_release_note(j, note_head_flag, is_note_head1, git_log_list)        
        note_text.append(release_note_text)#加入 release note

        wt_xlsx(memberName, note_text, git_log_list[date_flag][12:-1], team)

#將 git log 的 release note 合併成完整的一筆note
def make_release_note(j, note_head_flag, is_note_head, git_log_list):
    release_note_text = ''
        
    if j != len(note_head_flag) - 1:
        for k in range(note_head_flag[j] + 4, note_head_flag[j + 1]):
            if is_note_head(git_log_list, k):
                break
            if git_log_list[k] == '' or '-------------------------' in git_log_list[k]:
                continue
            release_note_text += git_log_list[k]
    else:
        for k in range(note_head_flag[j] + 4, len(git_log_list)):
            if is_note_head(git_log_list, k):
                break
            release_note_text += git_log_list[k]
    
    return release_note_text


def main():
    starttime = datetime.datetime.now() #開始時間戳記
    file_name=input("請輸入檔案名稱\n")
    merge_log_required=input("需要merge的檔案嗎？(y/n)\n")
    history_required=input("需要存入History資料表內嗎？(y/n)\n")
    
    #讀取要整理的 git log 資料表, 把每一行分開存進list資料格式中
    fin=open(file_name+'.txt', encoding='utf-8')
    git_log_list=fin.readlines()
    
    #找出每一筆 git log的開頭, 在list中的哪個位置
    i = 0
    note_head_flag1 = [] #記錄全部非 merge 的 log開頭位置
    note_head_flag2 = [] #記錄全部 merge 的 log開頭位置
    while True:
        if is_note_head1(git_log_list, i):
            note_head_flag1.append(i)

        if is_note_head2(git_log_list, i):
            note_head_flag2.append(i)

        i += 1

        if i >= len(git_log_list):
            break
    #把 log 寫入 excel
    add_note(git_log_list, note_head_flag1)
    if merge_log_required=='y':
        add_note2(git_log_list, note_head_flag2)
    
    print("正式 commit 共 " + str(len(note_head_flag1)) + " 筆")
    print("Merge commit 共 " + str(len(note_head_flag2)) + " 筆")
    
    #log 寫入 excel 完成, 將 excel 工作表儲存後關閉
    global workbook, history_workbook
    workbook.save('gitLog.xlsx')
    if history_required=="y":
        history_workbook.save('gitLogHistory.xlsx')
    
    endtime = datetime.datetime.now()#結束時間戳記
    print (endtime - starttime)      #總共運行時間



if __name__ == "__main__":
    main()
